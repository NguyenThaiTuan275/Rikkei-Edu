import openpyxl  # Thư viện xử lý Excel
from tkinter import messagebox
import os  # Để xử lý đường dẫn tệp

def export_to_excel(data):
    """
    Hàm xuất dữ liệu sang file Excel với tên cố định SinhVien.xlsx.

    Args:
        data (dict): Dữ liệu cần xuất gồm các thông tin như Tên, MSSV, Ngày sinh, Ngành học, Khoa, Khóa học.

    Returns:
        None: Xuất dữ liệu và lưu thành file Excel.
    """
    # Đường dẫn lưu file cố định trong cùng thư mục với chương trình
    file_path = os.path.join(os.getcwd(), "SinhVien.xlsx")

    try:
        # Tạo workbook và sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Thông tin thẻ sinh viên"

        # Ghi tiêu đề cột
        headers = ["Họ và tên", "MSSV", "Ngày sinh", "Ngành học", "Khoa", "Khóa học"]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        # Ghi dữ liệu từ dict vào Excel
        row_data = [
            data.get("Tên", ""),
            data.get("Mã sinh viên", ""),
            data.get("Ngày sinh", ""),
            data.get("Ngành học", ""),
            data.get("Khoa", ""),
            data.get("Khóa học", "")
        ]
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=2, column=col_num, value=value)

        # Lưu file Excel tự động với tên SinhVien.xlsx
        wb.save(file_path)
        messagebox.showinfo("Thành công", f"Dữ liệu đã được lưu vào file Excel: {file_path}")

    except Exception as e:
        messagebox.showerror("Lỗi", f"Không thể lưu file Excel: {e}")
